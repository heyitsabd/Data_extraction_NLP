from openpyxl import load_workbook
import requests
from bs4 import BeautifulSoup
import pandas as pd
from Data_read import Data_read

# Load the workbook and select the active sheet
wb = load_workbook('input.xlsx')
ws = wb.active

# Define the cell ranges
url_cells = [cell for row in ws['B2':'B' + str(ws.max_row)] for cell in row]
id_cells = [cell for row in ws['A2':'A' + str(ws.max_row)] for cell in row]

# Dictionary to store the data
data = {'URL_ID': [], 'WebData': []}

# Iterate over the URL and ID cells
for url_cell, id_cell in zip(url_cells, id_cells):
    url = url_cell.value
    url_id = id_cell.value
    try:
        response = requests.get(url)
        response.raise_for_status()
        soup = BeautifulSoup(response.text, 'html.parser')
        
        # Extract headings and main content
        web_contents = soup.select('div.td-ss-main-content')
        headings = soup.select('h1.entry-title')
        
        content = ''
        for heading in headings:
            if heading.string:
                content += heading.string + '\n'
        for web_content in web_contents:
            if web_content.text:
                content += web_content.text + '\n'
    
        # Save the content to a text file named after the URL ID
        with open(f'./Extracted_Data/{url_id}.txt', 'w', encoding='utf-8') as file:
            file.write(content)
        data['URL_ID'].append(url_id)
        data['WebData'].append(content)
    
    except requests.exceptions.RequestException as e:
        print(f"Error fetching {url}: {e}")


# Convert the dictionary to a DataFrame and save to an Excel file
# df = pd.DataFrame.from_dict(data)
# df.to_excel('result.xlsx', index=False)
