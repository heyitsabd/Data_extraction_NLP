from openpyxl import load_workbook
import os

# Load the workbooks and select the active worksheets
input_wb = load_workbook('input.xlsx')
input_ws = input_wb.active

output_wb = load_workbook('Output Data Structure.xlsx')
output_ws = output_wb.active

# Collect cell references for the IDs and result cells
id_cells = [cell for row in input_ws['A2':'A' + str(input_ws.max_row)] for cell in row]
result_positive_cells = [cell for row in output_ws['C2':'C' + str(output_ws.max_row)] for cell in row]
result_negative_cells = [cell for row in output_ws['D2':'D' + str(output_ws.max_row)] for cell in row]
polarity_score_cells = [cell for row in output_ws['E2':'E'+str(output_ws.max_row)] for cell in row]
subjetivity_score = [cell for row in output_ws['F2':'F'+str(output_wb.max_row)] for cell in row]

idx=0
for cell_positive_val,cell_negative_val in zip(result_positive_cells,result_negative_cells):
    polarity_score_cells[idx].value=(cell_positive_val.value-cell_negative_val.value)/((cell_positive_val.value+cell_negative_val.value)+0.000001)
    subjetivity_score[idx].value=(cell_positive_val+cell_negative_val)
    idx+=1
# Load positive and negative words
with open('positive-words.txt', 'r', encoding='utf-8') as word_file:
    positive_words = word_file.read().split()

with open('negative-words.txt', 'r', encoding="ISO-8859-1'") as negative_word_file:
    negative_words = negative_word_file.read().split()

def countWords(words_to_count, words):
    total_words = 0
    for word in words_to_count:
        total_words += words.count(word)
    return total_words

# Initialize dictionaries to store results
results_positive = {}
results_negative = {}

# Process each ID and calculate word counts
for id_cell in id_cells:
    url_id = id_cell.value
    try:
        with open(f'./Extracted_Data/{url_id}.txt', 'r', encoding='utf-8') as file:
            file_data = file.read()
            words = file_data.split()
            results_positive[url_id] = countWords(positive_words, words)
            results_negative[url_id] = countWords(negative_words, words)
    except Exception as e:
        print(f"Error processing {url_id}: {e}")

# Write positive word counts to the output sheet
for idx, id_cell in enumerate(id_cells):
    url_id = id_cell.value
    if url_id in results_positive:
        result_positive_cells[idx].value = results_positive[url_id]

# Write negative word counts to the output sheet
for idx, id_cell in enumerate(id_cells):
    url_id = id_cell.value
    if url_id in results_negative:
        result_negative_cells[idx].value = results_negative[url_id]

# Save the updated output Excel file
output_wb.save('Output Data Structure.xlsx')


print("Word counts have been successfully written to the output Excel file.")
